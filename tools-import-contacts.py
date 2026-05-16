"""
Hotel Contact List importer — re-runnable.

Reads `Hotel Contact List (Live).xlsx`, joins per-property leadership +
ownership + corporate-support assignments, and writes the result into
`data.contacts` in `data.json`.

Usage:
    # Dry-run — shows the join mapping and the diff vs current data.contacts
    python tools-import-contacts.py

    # Actually write
    python tools-import-contacts.py --apply

    # Override input file path
    python tools-import-contacts.py --xlsx "C:/some/other/contacts.xlsx" --apply

The hub does NOT need to be stopped during import — we use the same atomic
write pattern as the hub (write `data.json.tmp`, fsync, rename). Worst case
on a concurrent hub write: one of the two writes wins; nothing corrupts.
But: if the hub is running and you import, any unsaved hub-side mutation
between this script's load and rename can be lost. Recommend importing
during a quiet moment (no live editing in the dashboard).

Re-import strategy:
- The xlsx is the source of truth for per-property meta + leadership +
  corporate roster. Each apply-run REPLACES `data.contacts.property[<id>]`
  fields and `data.contacts.corporate` wholesale with what's in the sheet.
- `data.contacts.unmapped` holds sheet rows we couldn't join to a hub
  property — preserved on each run so you can see what's still drifting.
- `data.contacts._meta` records importedAt + sourceFile + schemaVersion.
- We touch nothing else in `data.json`.
"""

from __future__ import annotations
import argparse
import json
import os
import re
import sys
import tempfile
import time
from pathlib import Path
from typing import Any
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────────────────
# Paths
# ─────────────────────────────────────────────────────────────────────────────
HUB_ROOT = Path(__file__).resolve().parent
DATA_JSON = HUB_ROOT / "data.json"
DEFAULT_XLSX = Path("C:/Users/Owner/Downloads/Hotel Contact List (Live).xlsx")
SCHEMA_VERSION = 1

# ─────────────────────────────────────────────────────────────────────────────
# Hub property roster — kept in sync with server.js getPropertyList().
# Used as the join target for the spreadsheet's Hotel Name column.
# Update this list if a property is added / renamed in the hub.
# ─────────────────────────────────────────────────────────────────────────────
HUB_PROPERTIES = [
    (1,  "Embassy Suites Chicago Naperville",       "Hilton",   "IL", True,  False),
    (2,  "Hampton Inn Suites Chicago Schaumburg",   "Hilton",   "IL", True,  False),
    (3,  "Tru by Hilton Holland",                   "Hilton",   "MI", True,  False),
    (4,  "Home2 Suites Holland",                    "Hilton",   "MI", True,  False),
    (5,  "DoubleTree Winston Salem",                "Hilton",   "NC", True,  False),
    (6,  "Home2 Suites Normal",                     "Hilton",   "IL", True,  False),
    (7,  "Home2 Suites Fort Wayne",                 "Hilton",   "IN", False, True),  # coming-soon
    (8,  "Home2 Suites Plano",                      "Hilton",   "TX", True,  False),
    (9,  "TownePlace Suites Mesquite",              "Marriott", "TX", True,  False),
    (10, "Mainstay Suites Lexington",               "Choice",   "KY", True,  False),
    (11, "Quality Inn Lexington",                   "Choice",   "KY", True,  False),
    (12, "Home2 Suites Lexington Hamburg",          "Hilton",   "KY", True,  False),
    (13, "Home2 Suites Owensboro",                  "Hilton",   "KY", True,  False),
    (14, "TownePlace Suites Owensboro",             "Marriott", "KY", True,  False),
    (15, "Hilton Garden Inn Atlanta Airport North", "Hilton",   "GA", True,  False),
    (16, "Home2 Suites Evansville",                 "Hilton",   "IN", True,  False),
    (17, "Tru by Hilton Northlake",                 "Hilton",   "TX", True,  False),
    (18, "Holiday Inn Lexington",                   "IHG",      "KY", True,  False),
    (19, "Home2 Suites Prosper",                    "Hilton",   "TX", False, True),  # coming-soon
]


# ─────────────────────────────────────────────────────────────────────────────
# Normalization + fuzzy matching
# ─────────────────────────────────────────────────────────────────────────────
def normalize_name(s: Any) -> str:
    """Lowercase, strip punctuation/spaces/ampersands/parentheticals.

    Used both for hotel name matching and as a slug base. Aggressive enough
    that 'Home2Suites - Plano, TX', 'Home2 Suites - Plano', and 'Home2 Suites
    Plano' all produce 'home2suitesplano'.

    Also drops the word 'and' and the '&' character entirely so that
    'Hampton Inn & Suites - Chicago Schaumburg' (sheet) matches
    'Hampton Inn Suites Chicago Schaumburg' (hub).
    """
    s = "" if s is None else str(s)
    # Strip parenthetical suffixes like ' (1st Floor)' or ' (2nd & 3rd Floor)'
    s = re.sub(r"\([^)]*\)", "", s)
    s = s.lower()
    # Drop ampersand AND the word 'and' so they don't matter for matching
    s = s.replace("&", " ")
    s = re.sub(r"\band\b", " ", s)
    # 'home2suites' <-> 'home2 suites' — normalize digit-word boundaries
    s = re.sub(r"([a-z])(\d)", r"\1 \2", s)
    s = re.sub(r"(\d)([a-z])", r"\1 \2", s)
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def name_tokens(s: Any) -> set[str]:
    """Tokenize for overlap matching. Strips parens, ampersand, 'and',
    and short connector words. Returns a set of distinctive tokens."""
    s = "" if s is None else str(s)
    s = re.sub(r"\([^)]*\)", "", s)
    s = s.lower()
    s = s.replace("&", " ")
    # Split on non-alphanumeric, keep digit-word splits as separate tokens
    raw = re.split(r"[^a-z0-9]+", s)
    # Further split digit/letter boundaries inside tokens (home2suites -> home2 + suites)
    out = set()
    for tok in raw:
        if not tok:
            continue
        # Split mid-token if a digit follows letters or vice versa
        parts = re.findall(r"\d+|[a-z]+", tok)
        for p in parts:
            if len(p) >= 2 and p not in {"and", "the", "of", "by", "at", "in", "on", "tx", "il", "ky"}:
                out.add(p)
    return out


def slugify(s: Any) -> str:
    s = "" if s is None else str(s)
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def clean_phone(v: Any) -> str | None:
    """Normalize phone — strip floats from xlsx ('3367679595.0' -> '(336) 767-9595')."""
    if v in (None, ""):
        return None
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    digits = re.sub(r"\D", "", s)
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    if len(digits) == 11 and digits.startswith("1"):
        return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
    # Already formatted or non-standard — return cleaned but unformatted
    return s


def clean_zip(v: Any) -> str | None:
    if v in (None, ""):
        return None
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def clean_id(v: Any) -> str | None:
    """Numeric IDs in xlsx come through as floats — strip the .0 trail."""
    if v in (None, ""):
        return None
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def clean_str(v: Any) -> str | None:
    if v in (None, ""):
        return None
    s = str(v).strip()
    return s if s else None


def clean_email(v: Any) -> str | None:
    e = clean_str(v)
    if not e or "@" not in e:
        return None
    return e


# ─────────────────────────────────────────────────────────────────────────────
# Hotel name -> hub property id
# ─────────────────────────────────────────────────────────────────────────────
def build_hub_index() -> tuple[dict[str, int], dict[int, set[str]]]:
    """Returns (norm_to_pid, pid_to_tokens) for fuzzy matching."""
    norm_to_pid = {}
    pid_to_tokens = {}
    for pid, name, *_ in HUB_PROPERTIES:
        norm_to_pid[normalize_name(name)] = pid
        pid_to_tokens[pid] = name_tokens(name)
    return norm_to_pid, pid_to_tokens


# Manual-override mapping table — sheet 5-letter code -> hub propId.
# Use only for rows the fuzzy matcher gets wrong. Edit when a sheet row is
# added that doesn't auto-match its hub property.
MANUAL_CODE_OVERRIDES: dict[str, int] = {
    # (none currently — fuzzy matcher handles all known cases after the
    # ampersand + parens + token-overlap improvements)
}


def match_property(sheet_name: str, sheet_code: str | None,
                   norm_to_pid: dict[str, int],
                   pid_to_tokens: dict[int, set[str]]) -> int | None:
    """Resolution order:
       1. Manual code override (if MANUAL_CODE_OVERRIDES has this sheet's code)
       2. Exact normalized name match
       3. Substring containment (either direction)
       4. Token-overlap: at least 2 distinctive tokens shared, AND the sheet's
          tokens are a subset of (or equal to) the hub's tokens.
    """
    # 1. manual override
    if sheet_code and sheet_code in MANUAL_CODE_OVERRIDES:
        return MANUAL_CODE_OVERRIDES[sheet_code]
    if not sheet_name:
        return None
    # 2. exact normalized
    n = normalize_name(sheet_name)
    if n in norm_to_pid:
        return norm_to_pid[n]
    # 3. substring containment (lighter than token match — catches abbreviations)
    for hub_norm, pid in norm_to_pid.items():
        if hub_norm and (hub_norm in n or n in hub_norm):
            return pid
    # 4. STRICT subset match — sheet tokens must be a subset of hub tokens.
    #    Catches "Home2 Evansville" -> "Home2 Suites Evansville" (sheet tokens
    #    {home2, evansville} subset of hub {home2, suites, evansville}).
    #    REJECTS "Hampton Inn Denison TX" -> "Hampton Inn Suites Schaumburg"
    #    because {hampton, inn, denison} is NOT a subset of {hampton, inn,
    #    suites, chicago, schaumburg} ('denison' not in hub).
    #    Requires >=2 shared tokens to avoid trivial matches.
    sheet_tokens = name_tokens(sheet_name)
    if len(sheet_tokens) < 2:
        return None
    matches = []
    for pid, hub_tokens in pid_to_tokens.items():
        if sheet_tokens.issubset(hub_tokens) and len(sheet_tokens) >= 2:
            matches.append((pid, len(sheet_tokens & hub_tokens)))
    if not matches:
        return None
    # If multiple hubs satisfy the subset (rare), pick the one with the most
    # overlapping tokens — i.e. the hub closest in size to the sheet name.
    matches.sort(key=lambda x: -x[1])
    return matches[0][0]


# ─────────────────────────────────────────────────────────────────────────────
# Sheet parsers — one per worksheet
# ─────────────────────────────────────────────────────────────────────────────

# Sheet 1 — header on row 2; data starts row 3.
# Column letter -> internal field key (None = ignored).
SHEET1_COLS = {
    "A":  "hotelName",
    "B":  "brand",                # informational only — hub already knows brand
    "C":  "propertyCode",         # e.g. "WSM-DBT150"
    "D":  "code",                 # 5-letter PMS code, e.g. "INTWU" — joins to Ownership.D
    "E":  "rooms",
    "F":  "phone",
    "G":  "address",
    "H":  "city",
    "I":  "state",
    "J":  "zip",
    "K":  "supportController",
    "L":  "supportRegional",      # RDO
    "M":  "supportSalesRegional", # RSM
    "N":  "supportRevenue",
    "O":  "supportAreaOps",
    "P":  "gmName",
    "Q":  "gmEmail",
    "R":  "gmCell",
    "S":  "dosName",
    "T":  "dosEmail",
    "U":  "website",
    "V":  "strCode",
    "W":  "initialDate",
    "Y":  "linkedin",
    "Z":  "paylocityId",
    "AA": "shOrientation",
    "AB": "delphiId",
    "AC": "hotelEffectiveness",
    "AD": "m3CompanyCode",
    "AE": "m3PropertyCode",
    "AF": "otaBookings",
    "AG": "otaExpedia",
    "AH": "otaAgoda",
    "AI": "regDuns",
    "AJ": "regSamGov",
    "AK": "regFema",
    "AL": "regCage",
    "AM": "agmName",
    "AN": "agmEmail",
    "AO": "hotelEmailAP",
}

# Numeric/identifier columns that need .0 trimming (vs free-text)
SHEET1_NUMERIC_COLS = {
    "E", "F", "J", "V", "Z", "AF", "AG", "AH", "AI",
}


def col_letter_to_idx(letters: str) -> int:
    """A=0, B=1, ..., Z=25, AA=26, AB=27, ..."""
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def parse_sheet1(ws, norm_to_pid: dict[str, int], pid_to_tokens: dict[int, set[str]]
                 ) -> tuple[dict[int, dict], dict[str, dict]]:
    """Returns (mapped_by_propId, unmapped_by_code).

    Each mapped value is the per-property contacts dict ready for
    data.contacts.property[propId]. unmapped collects sheet rows we
    couldn't join to a hub property — preserved for visibility.
    """
    mapped: dict[int, dict] = {}
    unmapped: dict[str, dict] = {}

    rows = list(ws.iter_rows(values_only=True))
    # Header is row 2 (index 1); data starts row 3 (index 2)
    for ri in range(2, len(rows)):
        row = rows[ri]
        # Skip rows with no hotel name in column A
        a = row[0] if len(row) > 0 else None
        if not a:
            continue

        # Pull cells by configured column letters
        rec_raw: dict[str, Any] = {}
        for letters, key in SHEET1_COLS.items():
            idx = col_letter_to_idx(letters)
            v = row[idx] if idx < len(row) else None
            if letters in SHEET1_NUMERIC_COLS:
                v = clean_id(v)
            rec_raw[key] = v

        # Match to hub propId by hotel name + code (for manual overrides)
        sheet_code = clean_str(rec_raw["code"])
        pid = match_property(rec_raw["hotelName"], sheet_code, norm_to_pid, pid_to_tokens)

        # Build the structured contacts dict
        rec = {
            "hotelName":      clean_str(rec_raw["hotelName"]),
            "brand":          clean_str(rec_raw["brand"]),
            "propertyCode":   clean_str(rec_raw["propertyCode"]),
            "code":           clean_str(rec_raw["code"]),
            "rooms":          clean_id(rec_raw["rooms"]),
            "phone":          clean_phone(rec_raw["phone"]),
            "address":        clean_str(rec_raw["address"]),
            "city":           clean_str(rec_raw["city"]),
            "state":          clean_str(rec_raw["state"]),
            "zip":            clean_zip(rec_raw["zip"]),
            "website":        clean_str(rec_raw["website"]),
            "linkedin":       clean_str(rec_raw["linkedin"]),
            "strCode":        clean_id(rec_raw["strCode"]),
            "paylocityId":    clean_id(rec_raw["paylocityId"]),
            "m3CompanyCode":  clean_str(rec_raw["m3CompanyCode"]),
            "m3PropertyCode": clean_str(rec_raw["m3PropertyCode"]),
            "delphiId":       clean_str(rec_raw["delphiId"]),
            "hotelEffectiveness": clean_str(rec_raw["hotelEffectiveness"]),
            "shOrientation":  clean_str(rec_raw["shOrientation"]),
            "hotelEmailAP":   clean_email(rec_raw["hotelEmailAP"]),
            "ota": {
                "bookings": clean_id(rec_raw["otaBookings"]),
                "expedia":  clean_id(rec_raw["otaExpedia"]),
                "agoda":    clean_id(rec_raw["otaAgoda"]),
            },
            "regIds": {
                "duns":   clean_str(rec_raw["regDuns"]),
                "samGov": clean_str(rec_raw["regSamGov"]),
                "fema":   clean_str(rec_raw["regFema"]),
                "cage":   clean_str(rec_raw["regCage"]),
            },
            "gm": {
                "name":  clean_str(rec_raw["gmName"]),
                "email": clean_email(rec_raw["gmEmail"]),
                "cell":  clean_phone(rec_raw["gmCell"]),
            },
            "agm": {
                "name":  clean_str(rec_raw["agmName"]),
                "email": clean_email(rec_raw["agmEmail"]),
            },
            "dos": {
                "name":  clean_str(rec_raw["dosName"]),
                "email": clean_email(rec_raw["dosEmail"]),
            },
            "support": {
                "controller":     clean_str(rec_raw["supportController"]),
                "regional":       clean_str(rec_raw["supportRegional"]),
                "salesRegional":  clean_str(rec_raw["supportSalesRegional"]),
                "revenue":        clean_str(rec_raw["supportRevenue"]),
                "areaOps":        clean_str(rec_raw["supportAreaOps"]),
            },
            "owner": None,  # filled by parse_ownership join below
        }

        if pid is None:
            # Park the row under its 5-letter code (or hotel name if no code)
            key = rec["code"] or slugify(rec["hotelName"])
            unmapped[key] = rec
        elif pid in mapped:
            # Multiple sheet rows resolved to the same hub property. First wins
            # (the sheet is generally ordered with the active row first); later
            # rows go into unmapped under their own code so we can see them.
            print(f"  [conflict] hub#{pid} already matched to "
                  f"'{mapped[pid].get('hotelName')}' — also matched by "
                  f"'{rec.get('hotelName')}' (kept first match)")
            key = (rec["code"] or slugify(rec["hotelName"])) + ".dup"
            unmapped[key] = rec
        else:
            mapped[pid] = rec

    return mapped, unmapped


def parse_ownership(ws) -> dict[str, dict]:
    """Returns code -> ownership dict. Header row is row 2; data starts row 3."""
    rows = list(ws.iter_rows(values_only=True))
    out: dict[str, dict] = {}
    for ri in range(2, len(rows)):
        row = rows[ri]
        # Code is in column D (index 3)
        code = clean_str(row[3] if len(row) > 3 else None)
        if not code:
            continue
        out[code] = {
            "ein":         clean_str(row[1] if len(row) > 1 else None),
            "legalEntity": clean_str(row[2] if len(row) > 2 else None),
            "code":        code,
            "group":       clean_str(row[4] if len(row) > 4 else None),
            "contact":     clean_str(row[5] if len(row) > 5 else None),
            "dba":         clean_str(row[6] if len(row) > 6 else None),
            "state":       clean_str(row[7] if len(row) > 7 else None),
            "taxRate":     clean_str(row[8] if len(row) > 8 else None),
        }
    return out


def parse_corporate(ws) -> dict[str, dict]:
    """Sheet 2 — header row 1; data starts row 2.

    Column A is "Name, Title" combined. Split on first comma.
    """
    rows = list(ws.iter_rows(values_only=True))
    out: dict[str, dict] = {}
    for ri in range(1, len(rows)):
        row = rows[ri]
        a = clean_str(row[0] if len(row) > 0 else None)
        if not a:
            continue
        if "," in a:
            name, title = a.split(",", 1)
            name = name.strip()
            title = title.strip()
        else:
            name = a
            title = None
        slug = slugify(name)
        if not slug:
            continue
        out[slug] = {
            "name":       name,
            "title":      title,
            "department": clean_str(row[2] if len(row) > 2 else None),
            "cell":       clean_phone(row[1] if len(row) > 1 else None),
            "email":      clean_email(row[3] if len(row) > 3 else None),
            "linkedin":   clean_str(row[4] if len(row) > 4 else None),
        }
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Top-level: build the data.contacts blob from an xlsx file
# ─────────────────────────────────────────────────────────────────────────────
def build_contacts(xlsx_path: Path) -> dict[str, Any]:
    if not xlsx_path.exists():
        sys.exit(f"ERROR: xlsx not found at {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    sheets = {n.strip(): wb[n] for n in wb.sheetnames}

    # Sheet 1 has a trailing space in its name in the file we got — be tolerant
    sheet1 = next((wb[n] for n in wb.sheetnames if "Superhost Hotel Contact List" in n), None)
    sheet_corp = next((wb[n] for n in wb.sheetnames if "Corporate Contacts" in n), None)
    sheet_own = next((wb[n] for n in wb.sheetnames if "Ownership" == n.strip()), None)

    if sheet1 is None or sheet_corp is None or sheet_own is None:
        sys.exit(f"ERROR: missing required sheet. Found: {wb.sheetnames}")

    norm_to_pid, pid_to_tokens = build_hub_index()

    print(f"\n=== Reading {xlsx_path.name} ===")
    mapped, unmapped = parse_sheet1(sheet1, norm_to_pid, pid_to_tokens)
    ownership = parse_ownership(sheet_own)
    corporate = parse_corporate(sheet_corp)

    # Join ownership into mapped properties by 5-letter code
    join_hits = 0
    join_misses = []
    for pid, rec in mapped.items():
        code = rec.get("code")
        if code and code in ownership:
            rec["owner"] = ownership[code]
            join_hits += 1
        else:
            join_misses.append((pid, rec.get("hotelName"), code))

    contacts = {
        "property": {str(pid): rec for pid, rec in mapped.items()},
        "corporate": corporate,
        "unmapped": unmapped,
        "_meta": {
            "importedAt":     time.strftime("%Y-%m-%dT%H:%M:%S"),
            "sourceFile":     xlsx_path.name,
            "schemaVersion":  SCHEMA_VERSION,
            "stats": {
                "propertiesMapped":   len(mapped),
                "propertiesUnmapped": len(unmapped),
                "ownershipJoined":    join_hits,
                "ownershipMissed":    len(join_misses),
                "corporateContacts":  len(corporate),
            },
        },
    }

    # ── Mapping report ──────────────────────────────────────────────────────
    print(f"\nProperty join (sheet -> hub):")
    for pid, name, *_ in HUB_PROPERTIES:
        rec = mapped.get(pid)
        if rec:
            badge = "[OK]"
            sheet_name = rec.get("hotelName") or "?"
            extra = f"  GM: {rec['gm']['name'] or '-'}  Owner: {(rec['owner'] or {}).get('group') or '-'}"
            print(f"  {badge}  hub#{pid:>2}  {name[:38]:<38}  <-  {sheet_name[:38]:<38}{extra}")
        else:
            print(f"  [--]  hub#{pid:>2}  {name[:38]:<38}  <-  (no sheet match)")

    if unmapped:
        print(f"\nSheet rows NOT joined to any hub property ({len(unmapped)}):")
        for code, rec in unmapped.items():
            print(f"  -  code={code}  hotel={rec.get('hotelName')}")

    if join_misses:
        print(f"\nProperties without ownership join ({len(join_misses)}):")
        for pid, name, code in join_misses:
            print(f"  -  hub#{pid}  {name}  (code={code})")

    print(f"\nCorporate contacts parsed: {len(corporate)}")
    print(f"  sample: {next(iter(corporate.values()))}")

    return contacts


# ─────────────────────────────────────────────────────────────────────────────
# Bench sync — populate data.gmBench from data.contacts
# ─────────────────────────────────────────────────────────────────────────────
# Mirrors `defaultGMRecord(prop)` in server.js. Kept here so dry-run can
# show the diff without the hub running.
def default_gm_record(pid: int, name: str) -> dict:
    return {
        "propertyId": pid,
        "propertyName": name,
        "gmName": "",
        "gmEmail": "",
        "gmTenureAtProperty": "",
        "gmTenureWithCompany": "",
        "performance": "",
        "potential": "",
        "riskLevel": "GREEN",
        "riskReason": "",
        "successor": {"name": "", "role": "", "readiness": "EXPOSED"},
        "notes": "",
        "tags": [],
        "lastReviewedAt": None,
        "updatedAt": None,
    }


# Fields the importer is allowed to touch on an existing bench record.
# Everything else is user-owned (subjective assessment) and preserved.
CONTACT_DERIVED_FIELDS = {"gmName", "gmEmail"}


def sync_bench_from_contacts(existing_bench: dict, contacts_property: dict) -> tuple[dict, dict]:
    """Returns (updated_bench, dict-of-changes-made).

    Updates only the contact-derived fields. Preserves every subjective field
    (performance, potential, riskLevel, riskReason, tenure, notes, tags,
    lastReviewedAt). Never deletes user data.
    """
    changes: dict[str, tuple[Any, Any]] = {}

    for pid_str, contact_rec in contacts_property.items():
        pid = int(pid_str)
        # Find/create bench record
        bench = existing_bench.get(pid_str) or existing_bench.get(pid) or default_gm_record(
            pid, contact_rec.get("hotelName") or f"Property {pid}"
        )
        # Make sure the bench has the canonical hub property name (use contacts
        # hotelName since that's what the user sees on the cards too)
        if contact_rec.get("hotelName") and not bench.get("propertyName"):
            bench["propertyName"] = contact_rec["hotelName"]

        gm = contact_rec.get("gm") or {}
        agm = contact_rec.get("agm") or {}

        # GM name / email — contacts is source of truth
        new_gm_name = (gm.get("name") or "").strip()
        new_gm_email = (gm.get("email") or "").strip()
        old_gm_name = (bench.get("gmName") or "").strip()
        old_gm_email = (bench.get("gmEmail") or "").strip()

        if new_gm_name and new_gm_name != old_gm_name:
            changes.setdefault(pid, {})["gmName"] = (old_gm_name or "(empty)", new_gm_name)
            bench["gmName"] = new_gm_name
        if new_gm_email and new_gm_email != old_gm_email:
            changes.setdefault(pid, {})["gmEmail"] = (old_gm_email or "(empty)", new_gm_email)
            bench["gmEmail"] = new_gm_email

        # Successor — pull AGM as default successor IF the user hasn't set
        # someone else. We do NOT overwrite a user-set successor name.
        existing_succ = bench.get("successor") or {"name": "", "role": "", "readiness": "EXPOSED"}
        existing_succ_name = (existing_succ.get("name") or "").strip()
        new_agm_name = (agm.get("name") or "").strip()

        if new_agm_name:
            # Guard: don't make someone their own successor. Sheet sometimes has
            # the same name in GM and AGM columns when one person fills both
            # roles. Bench succession by definition is a different person.
            same_as_gm = (new_gm_name and new_agm_name.lower() == new_gm_name.lower())
            if same_as_gm:
                changes.setdefault(pid, {})["_warn_agm_same_as_gm"] = (new_agm_name, "(skipped)")
            else:
                # If no existing successor, OR existing successor was set by a
                # prior bench-sync (role == 'AGM'), we can refresh it from contacts.
                existing_role = (existing_succ.get("role") or "").upper()
                user_set = existing_succ_name and existing_role not in ("AGM", "")
                if not existing_succ_name or (existing_role == "AGM" and existing_succ_name != new_agm_name):
                    changes.setdefault(pid, {})["successor.name"] = (existing_succ_name or "(empty)", new_agm_name)
                    existing_succ["name"] = new_agm_name
                    existing_succ["role"] = "AGM"
                    # Don't touch readiness — that's user-owned ("EXPOSED" by default
                    # is fine; if user has set "BRIDGE" or "COVERED" we keep it)
                elif user_set:
                    # User has set a different non-AGM successor (e.g. an external
                    # candidate or the DOO) — leave it alone. No change.
                    pass

        bench["successor"] = existing_succ
        bench["updatedAt"] = time.strftime("%Y-%m-%dT%H:%M:%S")
        existing_bench[pid_str] = bench

    return existing_bench, changes


# ─────────────────────────────────────────────────────────────────────────────
# Atomic write to data.json — mirrors server.js saveData pattern
# ─────────────────────────────────────────────────────────────────────────────
def write_data_json(updates: dict[str, Any]) -> None:
    if not DATA_JSON.exists():
        sys.exit(f"ERROR: {DATA_JSON} not found")

    with open(DATA_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)

    for k, v in updates.items():
        data[k] = v

    # Compact format — server's recent atomic save also writes compact
    serialized = json.dumps(data, separators=(",", ":"), ensure_ascii=False)

    tmp = DATA_JSON.with_suffix(".json.tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(serialized)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, DATA_JSON)
    print(f"\n[OK] Wrote {DATA_JSON} ({len(serialized):,} bytes)")


# ─────────────────────────────────────────────────────────────────────────────
# CLI entry
# ─────────────────────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Path to xlsx (default: Downloads)")
    p.add_argument("--apply", action="store_true", help="Actually write to data.json (otherwise dry-run)")
    p.add_argument("--no-sync-bench", action="store_true",
                   help="Skip syncing data.gmBench from contacts (default: sync)")
    p.add_argument("--out", default=None, help="Optional: also write the contacts blob to this JSON file for inspection")
    args = p.parse_args()

    contacts = build_contacts(Path(args.xlsx))

    if args.out:
        Path(args.out).write_text(json.dumps(contacts, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"\n[OK] Wrote inspection JSON to {args.out}")

    # ── Bench sync preview ─────────────────────────────────────────────────
    bench_updates = None
    if not args.no_sync_bench:
        # Load current data.json so we can see existing bench records and
        # diff against them. (Even in dry-run we just read.)
        try:
            with open(DATA_JSON, "r", encoding="utf-8") as f:
                current = json.load(f)
        except Exception as e:
            print(f"\n[WARN] Could not read {DATA_JSON} for bench-sync preview: {e}")
            current = {}

        existing_bench = dict(current.get("gmBench") or {})
        bench_updates, changes = sync_bench_from_contacts(existing_bench, contacts["property"])

        print("\nBench sync preview (data.gmBench <- data.contacts):")
        if not changes:
            print("  (no changes — bench already in sync with contacts)")
        else:
            for pid, fields in sorted(changes.items()):
                pname = next((n for (i, n, *_) in HUB_PROPERTIES if i == pid), f"hub#{pid}")
                print(f"  hub#{pid:>2}  {pname[:40]:<40}")
                for fkey, (old, new) in fields.items():
                    print(f"           {fkey:<18} {old!r}  ->  {new!r}")
        print(f"  total properties touched: {len(changes)}")
        print(f"  preserved fields per record: performance, potential, riskLevel,")
        print(f"    riskReason, gmTenure*, notes, tags, lastReviewedAt,")
        print(f"    successor.readiness (and successor.name if user-set non-AGM)")

    # ── Apply ──────────────────────────────────────────────────────────────
    if args.apply:
        print("\n-> Applying to data.json (atomic write)...")
        updates = {"contacts": contacts}
        if bench_updates is not None:
            updates["gmBench"] = bench_updates
        write_data_json(updates)
        print("\nDone. Restart the hub to pick up new contacts API endpoints if the server")
        print("code is also being updated this session.")
    else:
        print("\n(dry-run only - re-run with --apply to write data.json)")


if __name__ == "__main__":
    main()
