"""Full column dump of the Hotel Contact List xlsx — every header, every
sample value across all sheets. Run from the hub root."""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

XLSX = Path(sys.argv[1] if len(sys.argv) > 1
            else "C:/Users/Owner/Downloads/Hotel Contact List (Live).xlsx")
wb = load_workbook(XLSX, data_only=True)

def fmt(v):
    if v is None: return ""
    s = str(v).strip()
    if len(s) > 60: s = s[:57] + "..."
    return s

# Best-guess header row for each sheet (Sheet 1 row 2, Ownership row 2,
# others row 1). Then list every column header + a 2-row sample.
HEADER_ROWS = {
    'Superhost Hotel Contact List - ': 2,
    'Corporate Contacts': 1,
    'Ownership': 2,
    'ID Number': 1,
}

for name in wb.sheetnames:
    ws = wb[name]
    hr = HEADER_ROWS.get(name, 1)
    print(f"\n=== {name!r}  (header row {hr}, max_col {ws.max_column}, max_row {ws.max_row}) ===")
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[hr - 1] if len(rows) >= hr else ()
    sample1 = rows[hr] if len(rows) > hr else ()
    sample2 = rows[hr + 1] if len(rows) > hr + 1 else ()
    for i, h in enumerate(headers):
        col = get_column_letter(i + 1)
        s1 = fmt(sample1[i] if i < len(sample1) else "")
        s2 = fmt(sample2[i] if i < len(sample2) else "")
        print(f"  {col:>3}  {fmt(h):<32}  ex1={s1!r:<40}  ex2={s2!r}")
