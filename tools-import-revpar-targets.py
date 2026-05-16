"""
Imports 2026 RevPAR Index monthly budgets from the user's Google Sheet export
into the SHAI Hub data store. Run after node server is up.

Source: C:\\Users\\Owner\\Downloads\\2026 RevPAR Targets.xlsx
Sheet:  '2026', section 'RevPAR INDEX TARGET' (rows 4-20)
Target: data.byPeriod['2026-MM'][propId].manual.revparIdxBud (per month)

Re-runnable. Uses /api/properties/bulk-save which merges into existing manual fields.
"""
import json
import urllib.request
import openpyxl
from pathlib import Path

XLSX_PATH = r'C:\Users\Owner\Downloads\2026 RevPAR Targets.xlsx'
HUB_BASE = 'http://localhost:3000'

# Maps sheet property name (cleaned) → SHAI Hub property ID
NAME_TO_ID = {
    'DoubleTree Winston-Salem':            5,
    'Embassy Suites Naperville':           1,
    'Hampton Inn & Suites Schaumburg':     2,
    'Hilton Garden Inn Atlanta':          15,
    'Holiday Inn Lexington':              18,
    'Home2 Suites Evansville':            16,
    'Home2 Suites Lexington/Hamburg':     12,
    'Home2 Suites Normal':                 6,
    'Home2 Suites Owensboro':             13,
    'Home2 Suites Plano':                  8,
    'Home2 Suites Holland':                4,
    'Mainstay Lexington':                 10,
    'Quality Inn Lexington':              11,
    'TownePlace Suites Mesquite':          9,
    'TownePlace Suites Owensboro':        14,
    'Tru Holland':                         3,
    'Tru Northlake':                      17,
}

# RevPAR INDEX TARGET section: rows 4-20 (17 properties), Column A holds property name
# Monthly Budget columns: Jan=3, Feb=5, Mar=7, Apr=11, May=13, Jun=15, Jul=19, Aug=21, Sep=23, Oct=27, Nov=29, Dec=31
MONTH_COLS = {1:3, 2:5, 3:7, 4:11, 5:13, 6:15, 7:19, 8:21, 9:23, 10:27, 11:29, 12:31}

def load_targets():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['2026']
    targets = {}  # {prop_id: {month_num: revparIdxBud}}
    unmapped = []
    for r in range(4, 21):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        name = str(name).strip()
        prop_id = NAME_TO_ID.get(name)
        if not prop_id:
            unmapped.append((r, name))
            continue
        per_month = {}
        for month_num, col in MONTH_COLS.items():
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            try:
                per_month[month_num] = round(float(v), 2)
            except (ValueError, TypeError):
                pass
        if per_month:
            targets[prop_id] = per_month
    return targets, unmapped

def post_bulk_save(period, entries):
    body = json.dumps({'entries': entries}).encode()
    req = urllib.request.Request(
        f'{HUB_BASE}/api/properties/bulk-save?period={period}',
        data=body,
        headers={'Content-Type': 'application/json'},
        method='POST',
    )
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read())

def main():
    targets, unmapped = load_targets()
    print(f'Loaded targets for {len(targets)}/17 properties.')
    if unmapped:
        print('UNMAPPED rows (skipped):')
        for r, n in unmapped:
            print(f'  row {r}: {n!r}')

    # Group by period
    by_period = {}  # {period: [{id, manual}, ...]}
    for prop_id, months in targets.items():
        for m, v in months.items():
            period = f'2026-{m:02d}'
            by_period.setdefault(period, []).append({
                'id': prop_id,
                'manual': {'revparIdxBud': v},
            })

    total_writes = 0
    for period in sorted(by_period.keys()):
        entries = by_period[period]
        result = post_bulk_save(period, entries)
        print(f'  {period}: wrote {result.get("count", len(entries))} properties')
        total_writes += len(entries)

    print(f'\nDONE. {total_writes} (property × month) targets imported.')

if __name__ == '__main__':
    main()
